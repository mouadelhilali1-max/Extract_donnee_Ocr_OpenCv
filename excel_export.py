"""Export the reconstructed CATIA tree and its OCR evidence safely.

The CSV and JSON exports are deliberately produced before the Excel workbook:
Excel is often left open while the extraction is rerun, whereas the two
machine-readable exports remain available in that situation.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable

import pandas as pd

from ocr_config import CSV_FILE, EXCEL_FILE, JSON_FILE


TREE_COLUMNS = (
    "line",
    "node_id",
    "text",
    "level",
    "parent_id",
    "parent",
    "full_path",
    "conf",
    "observations_count",
    "review_needed",
    "review_reason",
    "hierarchy_status",
    "segment",
)

REVIEW_COLUMNS = (
    "line",
    "node_id",
    "text",
    "corrected_text",
    "image",
    "capture_index",
    "top",
    "global_y",
    "level",
    "parent_id",
    "parent",
    "full_path",
    "conf",
    "ocr_method",
    "ocr_alternatives",
    "review_crop",
    "observations_count",
    "review_reason",
    "hierarchy_status",
    "segment",
)


class Exporter:
    """Write a clean tree sheet, complete OCR audit and review queue."""

    def __init__(
        self,
        dataframe: pd.DataFrame,
        *,
        excel_file: Path = EXCEL_FILE,
        csv_file: Path = CSV_FILE,
        json_file: Path = JSON_FILE,
    ):
        self.df = dataframe.copy()
        self.excel_file = Path(excel_file)
        self.csv_file = Path(csv_file)
        self.json_file = Path(json_file)

    @staticmethod
    def _available_columns(dataframe: pd.DataFrame, preferred: Iterable[str]) -> list[str]:
        return [column for column in preferred if column in dataframe.columns]

    def _tree_dataframe(self) -> pd.DataFrame:
        columns = self._available_columns(self.df, TREE_COLUMNS)
        # Compatibility with the former OCR-only dataframe: never export an
        # empty sheet merely because hierarchy fields are not present yet.
        return self.df.loc[:, columns].copy() if columns else self.df.copy()

    def _review_dataframe(self) -> pd.DataFrame:
        if "review_needed" not in self.df.columns:
            review = self.df.iloc[0:0].copy()
        else:
            values = self.df["review_needed"]
            if pd.api.types.is_bool_dtype(values):
                mask = values.fillna(False)
            else:
                mask = values.astype(str).str.strip().str.casefold().isin({"1", "true", "yes", "oui"})
            review = self.df.loc[mask].copy()

        columns = self._available_columns(review, REVIEW_COLUMNS)
        return review.loc[:, columns].copy() if columns else review

    @staticmethod
    def _format_sheet(worksheet) -> None:
        """Apply a compact, readable layout without scanning huge workbooks."""
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)

        worksheet.freeze_panes = "A2"
        if worksheet.max_column:
            last_column = get_column_letter(worksheet.max_column)
            worksheet.auto_filter.ref = f"A1:{last_column}{max(1, worksheet.max_row)}"

        # Inspect a bounded sample: a thousands-row full_path column should
        # not make the workbook excessively wide or slow to save.
        for column_index in range(1, worksheet.max_column + 1):
            column_letter = get_column_letter(column_index)
            header = worksheet.cell(row=1, column=column_index)
            header.fill = header_fill
            header.font = header_font
            header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            sample_values = [header.value]
            for row_index in range(2, min(worksheet.max_row, 251) + 1):
                sample_values.append(worksheet.cell(row=row_index, column=column_index).value)
            longest = max((len(str(value)) for value in sample_values if value is not None), default=8)

            # Paths and OCR alternatives need room; all other columns stay
            # compact.  The upper cap keeps the sheet usable on a laptop.
            name = str(header.value or "").casefold()
            if name in {"full_path", "ocr_alternatives"}:
                width = min(68, max(28, longest + 2))
            elif name in {"text", "parent", "review_reason"}:
                width = min(44, max(14, longest + 2))
            else:
                width = min(24, max(10, longest + 2))
            worksheet.column_dimensions[column_letter].width = width

        worksheet.row_dimensions[1].height = 30

    def _write_excel(self, destination: Path) -> Path:
        destination.parent.mkdir(parents=True, exist_ok=True)
        tree = self._tree_dataframe()
        review = self._review_dataframe()

        with pd.ExcelWriter(destination, engine="openpyxl") as writer:
            tree.to_excel(writer, sheet_name="CATIA_TREE", index=False)
            if not self.df.empty:
                self.df.to_excel(writer, sheet_name="OCR_AUDIT", index=False)
            # Keep this sheet even when nothing requires review: its emptiness
            # is a useful explicit result, and it remains ready for feedback.
            review.to_excel(writer, sheet_name="REVIEW", index=False)

            for worksheet in writer.sheets.values():
                self._format_sheet(worksheet)
        return destination

    def _fallback_excel_path(self) -> Path:
        self.excel_file.parent.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        suffix = self.excel_file.suffix or ".xlsx"
        base_name = f"{self.excel_file.stem}_{timestamp}"
        candidate = self.excel_file.parent / f"{base_name}{suffix}"
        index = 2
        while candidate.exists():
            candidate = self.excel_file.parent / f"{base_name}_{index}{suffix}"
            index += 1
        return candidate

    def export_excel(self) -> Path:
        """Write Excel, falling back to a timestamped workbook if it is open."""
        try:
            path = self._write_excel(self.excel_file)
            print(f"Excel sauvegarde : {path}")
            return path
        except PermissionError:
            fallback = self._fallback_excel_path()
            path = self._write_excel(fallback)
            print(
                "Excel principal indisponible (probablement ouvert). "
                f"Export de secours sauvegarde : {path}"
            )
            return path

    def export_csv(self) -> Path:
        self.csv_file.parent.mkdir(parents=True, exist_ok=True)
        self.df.to_csv(self.csv_file, index=False, encoding="utf-8-sig")
        print(f"CSV sauvegarde : {self.csv_file}")
        return self.csv_file

    def export_json(self) -> Path:
        self.json_file.parent.mkdir(parents=True, exist_ok=True)
        self.df.to_json(self.json_file, orient="records", indent=4, force_ascii=False)
        print(f"JSON sauvegarde : {self.json_file}")
        return self.json_file

    def export_all(self) -> dict[str, Path]:
        """Export durable data first; an open Excel window cannot stop a run."""
        csv_path = self.export_csv()
        json_path = self.export_json()
        excel_path = self.export_excel()
        return {"csv": csv_path, "json": json_path, "excel": excel_path}


if __name__ == "__main__":
    print("Utilisez Exporter(dataframe).export_all() depuis le pipeline OCR.")
