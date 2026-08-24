"""Direct CATIA functional-tolerance export.

This program is deliberately independent from :mod:`main`.  ``main.py``
continues to implement the existing screenshot/OCR workflow for the
``Résultat d'un ensemble d'annotations`` subtree.  This file follows a
different route: it attaches to the active CATIA document through the V5
Automation interface, walks the semantic model, reads FTA annotations, and
writes a separate workbook.

CATIA does not expose the left Specification Tree as one universal COM tree
in every workbench/release.  The reader therefore walks the documented model
collections (Bodies, HybridBodies, OrderedGeometricalSets, AnnotationSets,
and their children) and records what it could actually read.  A missing
semantic relationship is reported in ``REVIEW``; it is never guessed from a
series number.  An optional ``--tree-excel`` input can supply a previously
exported tree when a CATIA installation hides the visual REF/process groups
from Automation.

The program is read-only: it does not save, update, close, select, or modify
the CATIA document.

Example (run from the CATIA project directory)::

    .\\.venv\\Scripts\\python.exe .\\functional_tolerance_export.py

Optional document opening and diagnostic mode::

    .\\.venv\\Scripts\\python.exe .\\functional_tolerance_export.py \\
        --document "C:\\models\\piece.CATPart" --diagnose

The workbook is written below ``results/functional_tolerances/runs`` unless
``--output-dir`` is supplied.
"""

from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass, field
from datetime import datetime
import json
import math
from pathlib import Path
import re
import sys
import time
from typing import Any, Iterable, Iterator, Sequence


SCRIPT_VERSION = "1.0.2"

DEFAULT_MAX_DEPTH = 18
DEFAULT_MAX_NODES = 50_000
DEFAULT_MAX_COLLECTION_ITEMS = 20_000

# These are deliberately explicit.  Accessing every attribute of a CATIA
# COM proxy can invoke methods or trigger expensive lazy loading, so the
# reader only visits known, read-only model collections.
SAFE_CHILD_COLLECTIONS: tuple[str, ...] = (
    "Products",
    "Bodies",
    "HybridBodies",
    "OrderedGeometricalSets",
    "GeometricalSets",
    "HybridShapes",
    "HybridSketches",
    "AnnotationSets",
    "Annotations",
    "Captures",
    "TPSViews",
    "Views",
    "GeometricElements",
)

SAFE_LABEL_PROPERTIES: tuple[str, ...] = (
    "Name",
    "PartNumber",
    "InstanceName",
    "Nomenclature",
    "Label",
    "DisplayName",
    "Text",
)

SAFE_ANNOTATION_PROPERTIES: tuple[str, ...] = (
    "Name",
    "Text",
    "Description",
    "Type",
    "SuperType",
    "TPSStatus",
    "Value",
    "Tolerance",
    "ToleranceValue",
    "UpperTolerance",
    "LowerTolerance",
    "ToleranceType",
    "GeometricToleranceType",
    "Symbol",
    "Multiplicity",
    "Quantity",
    "DatumReferences",
    "Datums",
)

SAFE_ZERO_ARGUMENT_METHODS: tuple[str, ...] = (
    "GetText",
    "ToleranceZone",
    "AssociatedRefFrame",
    "SemanticGDT",
    "IsAToleranceZone",
    "IsAnAssociatedRefFrame",
    "IsAppliedOnMultipleEntities",
    "GetGeometricToleranceType",
    "GetToleranceZone",
    "GetAssociatedRefFrame",
    "GetSemanticGDT",
)

REF_PATTERN = re.compile(r"^\s*REF(?:\s|$|[|_\-])", re.IGNORECASE)
PROCESS_PATTERN = re.compile(r"^\s*(\d{2})\s*[-–—_]\s*\S", re.IGNORECASE)
SERIES_PATTERN = re.compile(r"(?<![A-Z0-9])(\d{2}[A-Z]\d{2})(?![A-Z0-9])", re.IGNORECASE)
MULTIPLICITY_PATTERN = re.compile(r"(?<![A-Z0-9])(\d+)\s*[x×](?![A-Z0-9])", re.IGNORECASE)
# Datum frames may contain compound references such as ``A | B-C | D-E``.
# Keep the frame text intact; do not reduce it to the REF branch name.
DATUM_PATTERN = re.compile(
    r"(?<![A-Z0-9])([A-Z](?:(?:\s*(?:\||/|;|,)\s*[A-Z](?:\s*-\s*[A-Z])?)|(?:\s*-\s*[A-Z])){1,})(?![A-Z0-9])"
)

# Common names used by the CAA FTA type enumeration.  Unknown values are
# retained verbatim in ``symbol_raw`` and routed to REVIEW instead of being
# silently mapped to a possibly wrong symbol.
SYMBOL_NAME_MAP: dict[str, str] = {
    "fta_trueposition": "Position",
    "fta_positionofasurface": "Position",
    "fta_positionofaline": "Position",
    "fta_patterntruepos": "Position (pattern)",
    "fta_localization": "Localisation",
    "fta_localisation": "Localisation",
    "fta_perpendicularity": "Perpendicularité",
    "fta_parallelism": "Parallélisme",
    "fta_angularity": "Angularité",
    "fta_flatness": "Planéité",
    "fta_straightness": "Rectitude",
    "fta_circularity": "Circularité",
    "fta_cylindricity": "Cylindricité",
    "fta_profileofaline": "Profil d'une ligne",
    "fta_profileofasurface": "Profil d'une surface",
    "fta_concentricity": "Concentricité",
    "fta_symmetry": "Symétrie",
    "fta_circularrunout": "Battement circulaire",
    "fta_totalrunout": "Battement total",
}


class FunctionalToleranceError(RuntimeError):
    """A user-facing error that leaves the existing workflow untouched."""


def _text(value: Any) -> str:
    """Convert a COM scalar to a stable, bounded display string."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "True" if value else "False"
    if isinstance(value, float):
        if not math.isfinite(value):
            return str(value)
        return f"{value:g}"
    if isinstance(value, (str, int)):
        return str(value).strip()
    if isinstance(value, (tuple, list)):
        return " | ".join(part for part in (_text(item) for item in value) if part)
    # COM objects often have a useful Name when converted to text, but avoid
    # exposing a huge repr in the workbook.
    try:
        result = str(value).strip()
    except Exception:
        return ""
    return result[:500] if result else ""


def _is_scalar(value: Any) -> bool:
    return value is None or isinstance(value, (str, int, float, bool))


def _scalar_text(value: Any, *, nested_properties: Sequence[str] = ("Text", "Name", "Value")) -> str:
    """Render a scalar or one documented scalar child of a CATIA object.

    FTA ``Annotation.Text`` is often a ``Text`` COM object rather than a
    Python string.  Its own ``Text`` property contains the actual callout;
    unwrapping that one level prevents the series code from disappearing.
    """
    if _is_scalar(value):
        return _clean(_text(value))
    for property_name in nested_properties:
        # Some CATIA/pywin32 bindings expose a documented read-only property
        # as a zero-argument COM method.  These three names are explicit,
        # scalar accessors, so using the safe reader keeps the label rather
        # than treating it as missing.
        nested, _, _ = _read_scalar_or_object(value, property_name)
        if _is_scalar(nested):
            rendered = _clean(_text(nested))
            if rendered:
                return rendered
    return ""


def _clean(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def _normalise_code(value: str) -> str:
    return re.sub(r"\s+", "", value or "").upper()


def _codes(value: str) -> list[str]:
    return [_normalise_code(match) for match in SERIES_PATTERN.findall(value or "")]


def _first_code(value: str) -> str:
    values = _codes(value)
    return values[0] if values else ""


def _is_ref(label: str) -> bool:
    return bool(REF_PATTERN.search(label or ""))


def _is_process(label: str) -> bool:
    return bool(PROCESS_PATTERN.search(label or "")) and not bool(SERIES_PATTERN.fullmatch((label or "").strip()))


def _is_series(label: str) -> bool:
    value = _clean(label)
    return bool(re.fullmatch(r"\d{2}[A-Z]\d{2}(?:\s*\([^)]*\))?", value, re.IGNORECASE))


def _symbol_from_type(value: str) -> str:
    key = re.sub(r"[^a-z0-9]", "", (value or "").casefold())
    if not key:
        return ""
    # First use the exact known enumeration, then conservative fragments.
    for enum_name, display in SYMBOL_NAME_MAP.items():
        enum_key = re.sub(r"[^a-z0-9]", "", enum_name.casefold())
        if key == enum_key:
            return display
    fragments = (
        ("trueposition", "Position"),
        ("position", "Position"),
        ("local", "Localisation"),
        ("perpendicular", "Perpendicularité"),
        ("parallel", "Parallélisme"),
        ("angular", "Angularité"),
        ("flatness", "Planéité"),
        ("straightness", "Rectitude"),
        ("circularity", "Circularité"),
        ("cylindricity", "Cylindricité"),
        ("profile", "Profil"),
        ("concentric", "Concentricité"),
        ("symmetry", "Symétrie"),
        ("runout", "Battement"),
    )
    for fragment, display in fragments:
        if fragment in key:
            return display
    return ""


def _exception_text(error: BaseException) -> str:
    message = _clean(str(error))
    return f"{type(error).__name__}: {message}"[:500]


def _is_com_dispatch_proxy(value: Any) -> bool:
    """Return true for a pywin32 COM object, including callable collections.

    A CATIA collection such as ``Documents`` can expose a default ``Item``
    member.  pywin32 consequently makes the *collection object* callable,
    even though ``CATIA.Documents`` is a read-only property and must not be
    invoked with ``Documents()``.  The private ``_oleobj_`` handle is the
    stable pywin32 marker for this kind of dispatch proxy.
    """
    if value is None:
        return False
    try:
        return getattr(value, "_oleobj_", None) is not None
    except Exception:
        return False


def _safe_property(obj: Any, name: str) -> tuple[Any, str]:
    """Read one property without enumerating arbitrary COM attributes."""
    if obj is None:
        return None, "owner is empty"
    try:
        value = getattr(obj, name)
    except Exception as error:
        return None, _exception_text(error)
    # Do not mistake a callable COM collection proxy (for example
    # ``CATIA.Documents``) for a Python method.  Its default Item member is
    # callable, but the collection itself is the desired property value.
    if callable(value) and not _is_com_dispatch_proxy(value):
        return None, f"{name} is a method"
    return value, ""


def _safe_call(obj: Any, name: str, *args: Any) -> tuple[Any, str]:
    if obj is None:
        return None, "owner is empty"
    try:
        method = getattr(obj, name)
    except Exception as error:
        return None, _exception_text(error)
    if not callable(method):
        return method, ""
    try:
        return method(*args), ""
    except Exception as error:
        return None, _exception_text(error)


def _read_scalar_or_object(obj: Any, name: str) -> tuple[Any, str, str]:
    """Read a property, or call it only when CATIA exposes it as a method."""
    value, error = _safe_property(obj, name)
    if value is not None and not (error and "method" in error):
        return value, "property", error
    called, call_error = _safe_call(obj, name)
    if called is not None:
        return called, "method", call_error
    return None, "", call_error or error


def _read_first_object(obj: Any, names: Sequence[str]) -> tuple[Any, str, str]:
    """Try equivalent read-only CATIA property/method names in order."""
    errors: list[str] = []
    for name in names:
        value, source, error = _read_scalar_or_object(obj, name)
        if value is not None:
            return value, source, error
        if error:
            errors.append(f"{name}: {error}")
    return None, "", " | ".join(errors)


def _com_type_name(obj: Any) -> str:
    if obj is None:
        return ""
    for property_name in ("TypeName", "Type", "ClassName"):
        value, _ = _safe_property(obj, property_name)
        if _is_scalar(value) and _text(value):
            return _text(value)
    try:
        return type(obj).__name__
    except Exception:
        return ""


def _label_for(obj: Any) -> tuple[str, dict[str, str], list[str]]:
    """Read a user label and retain all scalar evidence used to find it."""
    evidence: dict[str, str] = {}
    errors: list[str] = []
    selected = ""
    for name in SAFE_LABEL_PROPERTIES:
        value, source, error = _read_scalar_or_object(obj, name)
        if error and value is None:
            errors.append(f"{name}: {error}")
        rendered = _scalar_text(value)
        if rendered:
            evidence[name] = rendered
            if not selected:
                selected = rendered
    return selected, evidence, errors


def _iter_collection_value(collection: Any, *, limit: int, prefix: str = "Item") -> Iterator[tuple[int, Any, str]]:
    """Iterate a CATIA collection object (CATIA indexes collections at 1)."""
    if collection is None:
        return
    count_value, _, count_error = _read_scalar_or_object(collection, "Count")
    if count_value is None:
        # A few CATIA collections support Python iteration through _NewEnum;
        # use it only as a bounded fallback and never inspect arbitrary attrs.
        try:
            iterator = iter(collection)
        except Exception:
            return
        for index, item in enumerate(iterator, 1):
            if index > limit:
                break
            if item is not None:
                yield index, item, f"{prefix}[{index}]"
        return
    try:
        count = max(0, min(int(count_value), limit))
    except (TypeError, ValueError, OverflowError):
        return
    for index in range(1, count + 1):
        item, item_error = _safe_call(collection, "Item", index)
        if item is not None:
            yield index, item, f"{prefix}[{index}]"


def _iter_collection(owner: Any, name: str, *, limit: int) -> Iterator[tuple[int, Any, str]]:
    collection, _, error = _read_scalar_or_object(owner, name)
    if collection is None:
        return
    yield from _iter_collection_value(collection, limit=limit, prefix=name)


def _iter_collection_or_self(owner: Any, child_collection_name: str, *, limit: int) -> Iterator[tuple[int, Any, str]]:
    """Handle both CATIA wrappers and collection properties.

    ``Part.AnnotationSets`` is itself a collection in some releases, while
    ``Product.GetTechnologicalObject('CATAnnotationSets')`` can return a
    wrapper containing an ``AnnotationSets`` property.  The same variation
    exists for ``AnnotationSet.Annotations``.
    """
    nested, _, nested_error = _read_scalar_or_object(owner, child_collection_name)
    if nested is not None:
        nested_items = list(_iter_collection_value(nested, limit=limit, prefix=child_collection_name))
        if nested_items:
            yield from nested_items
            return
    yield from _iter_collection_value(owner, limit=limit, prefix=child_collection_name)


@dataclass
class TreeNode:
    node_id: str
    parent_id: str
    level: int
    text: str
    node_type: str
    collection_source: str
    collection_index: int
    source_com_path: str
    object_type: str
    object_ref: Any = field(repr=False, default=None)
    raw_properties: dict[str, str] = field(default_factory=dict)
    errors: list[str] = field(default_factory=list)
    children_count: int = 0


class CatiaModelReader:
    """Read the semantic CATIA model with bounded, read-only COM access."""

    def __init__(self, *, max_depth: int = DEFAULT_MAX_DEPTH, max_nodes: int = DEFAULT_MAX_NODES,
                 max_collection_items: int = DEFAULT_MAX_COLLECTION_ITEMS):
        self.max_depth = max_depth
        self.max_nodes = max_nodes
        self.max_collection_items = max_collection_items
        self.nodes: list[TreeNode] = []
        self.audit: list[dict[str, Any]] = []
        self._seen: set[int] = set()
        self._truncated = False

    def _record_audit(self, *, path: str, obj: Any, properties: dict[str, str], errors: Sequence[str]) -> None:
        self.audit.append({
            "source_com_path": path,
            "object_type": _com_type_name(obj),
            "properties": json.dumps(properties, ensure_ascii=False, sort_keys=True),
            "errors": " | ".join(errors),
        })

    def _walk(self, owner: Any, *, parent_id: str, level: int, path: str, via: str, index: int) -> None:
        if level > self.max_depth:
            self._truncated = True
            return
        if len(self.nodes) >= self.max_nodes:
            self._truncated = True
            return
        identity = id(owner)
        if identity in self._seen:
            return
        self._seen.add(identity)

        label, evidence, errors = _label_for(owner)
        if not label:
            label = _com_type_name(owner) or "(objet CATIA sans nom)"
        node_id = f"N{len(self.nodes) + 1:05d}"
        node_type = classify_node(label)
        node = TreeNode(
            node_id=node_id,
            parent_id=parent_id,
            level=level,
            text=label,
            node_type=node_type,
            collection_source=via,
            collection_index=index,
            source_com_path=path,
            object_type=_com_type_name(owner),
            object_ref=owner,
            raw_properties=evidence,
            errors=errors,
        )
        self.nodes.append(node)
        self._record_audit(path=path, obj=owner, properties=evidence, errors=errors)

        child_count = 0
        for collection_name in SAFE_CHILD_COLLECTIONS:
            for child_index, child, child_via in _iter_collection(owner, collection_name, limit=self.max_collection_items):
                if len(self.nodes) >= self.max_nodes:
                    self._truncated = True
                    break
                child_count += 1
                child_path = f"{path}.{child_via}" if path else child_via
                before = len(self.nodes)
                self._walk(
                    child,
                    parent_id=node_id,
                    level=level + 1,
                    path=child_path,
                    via=collection_name,
                    index=child_index,
                )
                # The seen-set may suppress aliases; only count actual nodes.
                if len(self.nodes) == before:
                    child_count -= 1
            if self._truncated:
                break
        node.children_count = child_count

    def read(self, document: Any) -> list[TreeNode]:
        root = None
        root_name = ""
        for candidate_name in ("Part", "Product"):
            candidate, _, _ = _read_scalar_or_object(document, candidate_name)
            if candidate is not None:
                root = candidate
                root_name = candidate_name
                break
        if root is None:
            raise FunctionalToleranceError(
                "Le document CATIA actif ne fournit ni Part ni Product accessible par Automation."
            )
        self._walk(root, parent_id="", level=0, path=root_name, via="DOCUMENT", index=1)
        return self.nodes


def classify_node(label: str) -> str:
    if _is_ref(label):
        return "REF"
    if _is_series(label):
        return "SERIES"
    if _is_process(label):
        return "PROCESS"
    lower = (label or "").casefold()
    if "annotation" in lower or "tolér" in lower or "toler" in lower:
        return "ANNOTATION"
    return "OTHER"


@dataclass
class AnnotationRecord:
    record_id: str
    set_name: str
    annotation_index: int
    name: str
    text: str
    type_raw: str
    supertype_raw: str
    symbol: str
    symbol_raw: str
    tolerance_display: str
    tolerance_value: str
    tolerance_unit: str
    quantity_display: str
    quantity_count: str
    datum_references: str
    raw_callout: str
    source_method: str
    status: str
    review_needed: str
    review_reason: str
    raw_fields: dict[str, str] = field(default_factory=dict)
    audit_rows: list[dict[str, str]] = field(default_factory=list)


class FtaReader:
    """Read FTA annotations and preserve every raw COM value."""

    def __init__(self):
        self.records: list[AnnotationRecord] = []

    @staticmethod
    def _read_field(obj: Any, property_name: str, *, audit: list[dict[str, str]], record_id: str) -> str:
        value, source, error = _read_scalar_or_object(obj, property_name)
        rendered = _scalar_text(value)
        audit.append({
            "record_id": record_id,
            "field": property_name,
            "raw_value": rendered,
            "normalised_value": rendered,
            "CATIA_property": property_name,
            "source_method": "COM_DIRECT" if rendered else "COM_DIRECT_EMPTY",
            "evidence": source,
            "confidence": "high" if rendered else "unknown",
            "validation_status": "read" if rendered else "not_available",
            "error": error,
        })
        return rendered

    @staticmethod
    def _annotation_sets(document: Any) -> tuple[Any, str, list[str]]:
        errors: list[str] = []
        part, _, part_error = _read_scalar_or_object(document, "Part")
        if part is not None:
            sets, _, set_error = _read_scalar_or_object(part, "AnnotationSets")
            if sets is not None:
                return sets, "Part.AnnotationSets", errors
            if set_error:
                errors.append(f"Part.AnnotationSets: {set_error}")

        product, _, product_error = _read_scalar_or_object(document, "Product")
        if product is not None:
            technological, tech_error = _safe_call(product, "GetTechnologicalObject", "CATAnnotationSets")
            if technological is not None:
                # Loading a list is a read operation used by CATIA for CGR
                # documents.  It is deliberately optional and best effort.
                _safe_call(technological, "LoadAnnotationSetsList")
                return technological, "Product.GetTechnologicalObject(CATAnnotationSets)", errors
            if tech_error:
                errors.append(f"Product.CATAnnotationSets: {tech_error}")
        if product_error:
            errors.append(f"Product: {product_error}")
        return None, "", errors

    @staticmethod
    def _datum_text(ann: Any, *, raw_fields: dict[str, str], audit: list[dict[str, str]], record_id: str) -> str:
        is_frame, _, _ = _read_first_object(ann, ("IsAnAssociatedRefFrame", "IsAssociatedRefFrame"))
        frame_obj = None
        if bool(is_frame):
            frame_obj, _, frame_error = _read_first_object(ann, ("AssociatedRefFrame", "GetAssociatedRefFrame"))
            if frame_error:
                audit.append({
                    "record_id": record_id,
                    "field": "AssociatedRefFrame",
                    "raw_value": "",
                    "normalised_value": "",
                    "CATIA_property": "AssociatedRefFrame",
                    "source_method": "COM_DIRECT",
                    "evidence": "",
                    "confidence": "unknown",
                    "validation_status": "not_available",
                    "error": frame_error,
                })
        candidates: list[str] = []
        frame_values: list[str] = []
        for owner, prefix in ((ann, "Annotation"), (frame_obj, "ReferenceFrame")):
            if owner is None:
                continue
            for prop in ("DatumReferences", "Datums", "Text", "Name", "ReferenceFrame", "ReferenceFrame2"):
                value, source, error = _read_scalar_or_object(owner, prop)
                rendered = _scalar_text(value)
                if rendered:
                    candidates.append(rendered)
                    raw_fields[f"{prefix}.{prop}"] = rendered
                    audit.append({
                        "record_id": record_id,
                        "field": "datum_references",
                        "raw_value": rendered,
                        "normalised_value": rendered,
                        "CATIA_property": f"{prefix}.{prop}",
                        "source_method": "COM_DIRECT",
                        "evidence": source,
                        "confidence": "high",
                        "validation_status": "read",
                        "error": error,
                    })
            # ReferenceFrame.Frame(i) is exposed by some V5 releases.  Its
            # documented signature is ``Frame(first, second, third)`` with
            # three output strings.  pywin32 returns those output arguments
            # as a tuple when makepy information is available.  Try that
            # form first, then retain the bounded compatibility probes below.
            if prefix == "ReferenceFrame":
                value, error = _safe_call(owner, "Frame", "", "", "")
                if isinstance(value, (tuple, list)):
                    for frame_value in value[:3]:
                        rendered = _scalar_text(frame_value)
                        if rendered:
                            frame_values.append(rendered)
                            candidates.append(rendered)
                            raw_fields["ReferenceFrame.Frame"] = " | ".join(frame_values)
                            audit.append({
                                "record_id": record_id,
                                "field": "datum_references",
                                "raw_value": rendered,
                                "normalised_value": rendered,
                                "CATIA_property": "ReferenceFrame.Frame",
                                "source_method": "COM_DIRECT",
                                "evidence": "method_out_parameters",
                                "confidence": "high",
                                "validation_status": "read",
                                "error": error,
                            })
                elif _scalar_text(value):
                    rendered = _scalar_text(value)
                    candidates.append(rendered)
                    raw_fields["ReferenceFrame.Frame"] = rendered
                    audit.append({
                        "record_id": record_id,
                        "field": "datum_references",
                        "raw_value": rendered,
                        "normalised_value": rendered,
                        "CATIA_property": "ReferenceFrame.Frame",
                        "source_method": "COM_DIRECT",
                        "evidence": "method_return",
                        "confidence": "medium",
                        "validation_status": "read",
                        "error": error,
                    })
                for frame_index in (1, 2, 3, 0):
                    value, error = _safe_call(owner, "Frame", frame_index)
                    rendered = _scalar_text(value)
                    if rendered:
                        candidates.append(rendered)
                        if frame_index in (1, 2, 3):
                            frame_values.append(rendered)
                        raw_fields[f"ReferenceFrame.Frame[{frame_index}]"] = rendered
                        audit.append({
                            "record_id": record_id,
                            "field": "datum_references",
                            "raw_value": rendered,
                            "normalised_value": rendered,
                            "CATIA_property": f"ReferenceFrame.Frame[{frame_index}]",
                            "source_method": "COM_DIRECT",
                            "evidence": "method",
                            "confidence": "high",
                            "validation_status": "read",
                            "error": error,
                        })
        # CATIA's ReferenceFrame.Frame(1..3) represents the three datum
        # compartments.  Keeping those values in their native order is more
        # reliable than trying to infer a frame from arbitrary annotation
        # text.  Remove duplicate 0/1 based probes while preserving order.
        ordered_frame_values = list(dict.fromkeys(frame_values))
        if ordered_frame_values:
            return " | ".join(ordered_frame_values)
        # Prefer a string which visibly contains several datum letters.  Do
        # not synthesize a datum from the REF parent; that is a different
        # field in the Excel output.
        for candidate in candidates:
            match = DATUM_PATTERN.search(candidate.upper())
            if match:
                return _clean(match.group(1))
        # A label such as ``06B01 2X`` is useful audit evidence but is not a
        # datum reference.  Leave the value blank rather than exporting that
        # label as a false datum frame.
        return ""

    def _read_annotation(self, ann: Any, *, set_name: str, index: int) -> AnnotationRecord:
        record_id = f"A{len(self.records) + 1:05d}"
        audit: list[dict[str, str]] = []
        raw_fields: dict[str, str] = {}
        values: dict[str, str] = {}
        for prop in SAFE_ANNOTATION_PROPERTIES:
            value = self._read_field(ann, prop, audit=audit, record_id=record_id)
            if value:
                values[prop] = value
                raw_fields[prop] = value

        for method_name in SAFE_ZERO_ARGUMENT_METHODS:
            if method_name in values:
                continue
            value, source, error = _read_scalar_or_object(ann, method_name)
            rendered = _scalar_text(value)
            if rendered:
                values[method_name] = rendered
                raw_fields[method_name] = rendered
                audit.append({
                    "record_id": record_id,
                    "field": method_name,
                    "raw_value": rendered,
                    "normalised_value": rendered,
                    "CATIA_property": method_name,
                    "source_method": "COM_DIRECT",
                    "evidence": source,
                    "confidence": "high",
                    "validation_status": "read",
                    "error": error,
                })

        zone_obj, _, zone_error = _read_first_object(ann, ("ToleranceZone", "GetToleranceZone"))
        if zone_obj is not None and not _is_scalar(zone_obj):
            for prop in ("Value", "Form", "UpperTolerance", "LowerTolerance", "Unit"):
                value, source, error = _read_scalar_or_object(zone_obj, prop)
                rendered = _scalar_text(value)
                if rendered:
                    raw_fields[f"ToleranceZone.{prop}"] = rendered
                    values[f"ToleranceZone.{prop}"] = rendered
                    audit.append({
                        "record_id": record_id,
                        "field": f"ToleranceZone.{prop}",
                        "raw_value": rendered,
                        "normalised_value": rendered,
                        "CATIA_property": f"ToleranceZone.{prop}",
                        "source_method": "COM_DIRECT",
                        "evidence": source,
                        "confidence": "high",
                        "validation_status": "read",
                        "error": error,
                    })
        elif zone_error:
            audit.append({
                "record_id": record_id,
                "field": "ToleranceZone",
                "raw_value": "",
                "normalised_value": "",
                "CATIA_property": "ToleranceZone",
                "source_method": "COM_DIRECT",
                "evidence": "",
                "confidence": "unknown",
                "validation_status": "not_available",
                "error": zone_error,
            })

        # Newer CATIA releases expose the semantic GDT object separately.
        # NxDisplay is the authoritative source for a multiplicity such as
        # ``2X`` when it is available; it is intentionally read as evidence,
        # never reconstructed from the series number.
        semantic_obj, _, semantic_error = _read_first_object(ann, ("SemanticGDT", "GetSemanticGDT"))
        if semantic_obj is not None and not _is_scalar(semantic_obj):
            for prop in ("NxDisplay", "Quantity", "Multiplicity", "Text"):
                value, source, error = _read_scalar_or_object(semantic_obj, prop)
                rendered = _scalar_text(value)
                if rendered:
                    raw_fields[f"SemanticGDT.{prop}"] = rendered
                    values[f"SemanticGDT.{prop}"] = rendered
                    audit.append({
                        "record_id": record_id,
                        "field": f"SemanticGDT.{prop}",
                        "raw_value": rendered,
                        "normalised_value": rendered,
                        "CATIA_property": f"SemanticGDT.{prop}",
                        "source_method": "COM_DIRECT",
                        "evidence": source,
                        "confidence": "high",
                        "validation_status": "read",
                        "error": error,
                    })
            nx_obj, _, nx_error = _read_first_object(semantic_obj, ("NxDisplay",))
            if nx_obj is not None and not _is_scalar(nx_obj):
                for prop in ("Count", "Number", "Nx", "InstanceCount", "Type", "Text", "Name", "Value"):
                    value, source, error = _read_scalar_or_object(nx_obj, prop)
                    rendered = _scalar_text(value)
                    if rendered:
                        raw_fields[f"SemanticGDT.NxDisplay.{prop}"] = rendered
                        values[f"SemanticGDT.NxDisplay.{prop}"] = rendered
                        audit.append({
                            "record_id": record_id,
                            "field": f"SemanticGDT.NxDisplay.{prop}",
                            "raw_value": rendered,
                            "normalised_value": rendered,
                            "CATIA_property": f"SemanticGDT.NxDisplay.{prop}",
                            "source_method": "COM_DIRECT",
                            "evidence": source,
                            "confidence": "high",
                            "validation_status": "read",
                            "error": error,
                        })
                for method_name in ("IsACollection", "IsASeparate"):
                    value, source, error = _read_scalar_or_object(nx_obj, method_name)
                    rendered = _scalar_text(value)
                    if rendered:
                        raw_fields[f"SemanticGDT.NxDisplay.{method_name}"] = rendered
                        values[f"SemanticGDT.NxDisplay.{method_name}"] = rendered
                        audit.append({
                            "record_id": record_id,
                            "field": f"SemanticGDT.NxDisplay.{method_name}",
                            "raw_value": rendered,
                            "normalised_value": rendered,
                            "CATIA_property": f"SemanticGDT.NxDisplay.{method_name}",
                            "source_method": "COM_DIRECT",
                            "evidence": source,
                            "confidence": "high",
                            "validation_status": "read",
                            "error": error,
                        })
            elif nx_error:
                audit.append({
                    "record_id": record_id,
                    "field": "SemanticGDT.NxDisplay",
                    "raw_value": "",
                    "normalised_value": "",
                    "CATIA_property": "SemanticGDT.NxDisplay",
                    "source_method": "COM_DIRECT",
                    "evidence": "",
                    "confidence": "unknown",
                    "validation_status": "not_available",
                    "error": nx_error,
                })
            for method_name in ("IsAppliedOnMultipleEntities",):
                value, source, error = _read_scalar_or_object(semantic_obj, method_name)
                rendered = _scalar_text(value)
                if rendered:
                    raw_fields[f"SemanticGDT.{method_name}"] = rendered
                    values[f"SemanticGDT.{method_name}"] = rendered
                    audit.append({
                        "record_id": record_id,
                        "field": f"SemanticGDT.{method_name}",
                        "raw_value": rendered,
                        "normalised_value": rendered,
                        "CATIA_property": f"SemanticGDT.{method_name}",
                        "source_method": "COM_DIRECT",
                        "evidence": source,
                        "confidence": "high",
                        "validation_status": "read",
                        "error": error,
                    })
        elif semantic_error:
            audit.append({
                "record_id": record_id,
                "field": "SemanticGDT",
                "raw_value": "",
                "normalised_value": "",
                "CATIA_property": "SemanticGDT",
                "source_method": "COM_DIRECT",
                "evidence": "",
                "confidence": "unknown",
                "validation_status": "not_available",
                "error": semantic_error,
            })

        name = values.get("Name", "")
        text_value = values.get("Text", "") or values.get("GetText", "")
        type_raw = values.get("GeometricToleranceType", "") or values.get("Type", "")
        supertype_raw = values.get("SuperType", "")
        combined = " | ".join(value for value in (name, text_value, *values.values()) if value)
        symbol_raw = type_raw or values.get("Symbol", "")
        symbol = values.get("Symbol", "") or _symbol_from_type(symbol_raw)

        tolerance_display = (
            values.get("ToleranceZone.Value", "")
            or values.get("ToleranceValue", "")
            or values.get("Tolerance", "")
            or values.get("Value", "")
        )
        tolerance_value = tolerance_display
        tolerance_unit = values.get("ToleranceZone.Unit", "") or values.get("Unit", "")

        quantity_display = (
            values.get("Multiplicity", "")
            or values.get("Quantity", "")
            or values.get("SemanticGDT.NxDisplay", "")
            or values.get("SemanticGDT.NxDisplay.Count", "")
            or values.get("SemanticGDT.NxDisplay.Number", "")
            or values.get("SemanticGDT.NxDisplay.Nx", "")
            or values.get("SemanticGDT.NxDisplay.InstanceCount", "")
        )
        quantity_count = ""
        nx_count = (
            values.get("SemanticGDT.NxDisplay.Count", "")
            or values.get("SemanticGDT.NxDisplay.Number", "")
            or values.get("SemanticGDT.NxDisplay.Nx", "")
            or values.get("SemanticGDT.NxDisplay.InstanceCount", "")
        )
        if nx_count and not MULTIPLICITY_PATTERN.search(quantity_display):
            try:
                numeric_count = int(float(nx_count))
            except (TypeError, ValueError):
                numeric_count = 0
            if numeric_count > 0:
                quantity_display = f"{numeric_count}X"
                quantity_count = str(numeric_count)
        if not quantity_display:
            if values.get("SemanticGDT.NxDisplay.IsACollection", "").casefold() == "true":
                quantity_display = "Collection"
            elif values.get("SemanticGDT.NxDisplay.IsASeparate", "").casefold() == "true":
                quantity_display = "Separate"
            elif values.get("SemanticGDT.IsAppliedOnMultipleEntities", "").casefold() == "true":
                quantity_display = "multiple"
        quantity_match = MULTIPLICITY_PATTERN.search(combined)
        if quantity_match:
            quantity_display = quantity_match.group(0).replace("×", "X")
            quantity_count = quantity_match.group(1)
        elif quantity_display:
            quantity_match = MULTIPLICITY_PATTERN.search(quantity_display)
            if quantity_match:
                quantity_display = quantity_match.group(0).replace("×", "X")
                quantity_count = quantity_match.group(1)
        if not quantity_display:
            applied, _, _ = _read_scalar_or_object(ann, "IsAppliedOnMultipleEntities")
            if applied is not None:
                quantity_display = "multiple" if bool(applied) else "1X"
                quantity_count = "" if bool(applied) else "1"

        datum_references = self._datum_text(ann, raw_fields=raw_fields, audit=audit, record_id=record_id)
        status = "COM_DIRECT"
        review_reasons: list[str] = []
        if not name and not text_value:
            review_reasons.append("annotation_without_name_or_text")
        if not symbol:
            review_reasons.append("symbol_not_exposed")
        if not tolerance_display:
            review_reasons.append("tolerance_not_exposed")
        if not quantity_display:
            review_reasons.append("quantity_not_exposed")
        if not datum_references:
            review_reasons.append("datum_references_not_exposed")
        if review_reasons:
            status = "NEEDS_REVIEW"

        record = AnnotationRecord(
            record_id=record_id,
            set_name=set_name,
            annotation_index=index,
            name=name,
            text=text_value,
            type_raw=type_raw,
            supertype_raw=supertype_raw,
            symbol=symbol,
            symbol_raw=symbol_raw,
            tolerance_display=tolerance_display,
            tolerance_value=tolerance_value,
            tolerance_unit=tolerance_unit,
            quantity_display=quantity_display,
            quantity_count=quantity_count,
            datum_references=datum_references,
            raw_callout=combined[:2_000],
            source_method="COM_DIRECT",
            status=status,
            review_needed="Oui" if review_reasons else "Non",
            review_reason="; ".join(review_reasons),
            raw_fields=raw_fields,
            audit_rows=audit,
        )
        self.records.append(record)
        return record

    def read(self, document: Any) -> tuple[list[AnnotationRecord], list[dict[str, str]]]:
        sets, source, errors = self._annotation_sets(document)
        if sets is None:
            return [], [{
                "record_id": "",
                "field": "AnnotationSets",
                "raw_value": "",
                "normalised_value": "",
                "CATIA_property": source or "Part.AnnotationSets",
                "source_method": "COM_DIRECT",
                "evidence": "",
                "confidence": "unknown",
                "validation_status": "not_available",
                "error": " | ".join(errors) or "AnnotationSets inaccessible",
            }]
        _safe_call(sets, "LoadAnnotationSetsList")
        set_items = list(_iter_collection_or_self(sets, "AnnotationSets", limit=DEFAULT_MAX_COLLECTION_ITEMS))
        # If CATIA returned one AnnotationSet object rather than a collection,
        # treat it as a single set instead of silently returning no records.
        if not set_items:
            annotations_probe, _, _ = _read_scalar_or_object(sets, "Annotations")
            if annotations_probe is not None:
                set_items = [(1, sets, "AnnotationSets[1]")]
        for set_index, annotation_set, _ in set_items:
            # Some CATIA versions return the collection itself from the
            # technological object, while others return a wrapper containing
            # AnnotationSets.  Support both forms.
            set_name, _, _ = _label_for(annotation_set)
            annotations, _, _ = _read_scalar_or_object(annotation_set, "Annotations")
            if annotations is None and set_index == 1:
                annotations = sets
                set_name = source
            if annotations is None:
                continue
            for index, annotation, _ in _iter_collection_or_self(annotations, "Annotations", limit=DEFAULT_MAX_COLLECTION_ITEMS):
                self._read_annotation(annotation, set_name=set_name or f"AnnotationSet {set_index}", index=index)
        return self.records, [row for record in self.records for row in record.audit_rows]


def _ancestor_map(nodes: Sequence[TreeNode]) -> dict[str, TreeNode]:
    return {node.node_id: node for node in nodes}


def _path_for(node: TreeNode, by_id: dict[str, TreeNode]) -> list[TreeNode]:
    path: list[TreeNode] = []
    current: TreeNode | None = node
    seen: set[str] = set()
    while current is not None and current.node_id not in seen:
        seen.add(current.node_id)
        path.append(current)
        current = by_id.get(current.parent_id)
    path.reverse()
    return path


def _node_order(node: TreeNode, siblings: dict[str, list[TreeNode]]) -> int:
    values = siblings.get(node.parent_id, [])
    for position, sibling in enumerate(values, 1):
        if sibling.node_id == node.node_id:
            return position
    return node.collection_index or 0


def _build_siblings(nodes: Sequence[TreeNode]) -> dict[str, list[TreeNode]]:
    result: dict[str, list[TreeNode]] = {}
    for node in nodes:
        result.setdefault(node.parent_id, []).append(node)
    return result


def _annotation_index(records: Sequence[AnnotationRecord]) -> dict[str, list[AnnotationRecord]]:
    result: dict[str, list[AnnotationRecord]] = {}
    for record in records:
        source = " | ".join((record.name, record.text, record.raw_callout))
        for code in dict.fromkeys(_codes(source)):
            result.setdefault(code, []).append(record)
    return result


def build_tolerance_rows(nodes: Sequence[TreeNode], records: Sequence[AnnotationRecord]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    by_id = _ancestor_map(nodes)
    siblings = _build_siblings(nodes)
    ann_by_code = _annotation_index(records)
    ref_nodes = [node for node in nodes if node.node_type == "REF"]
    rows: list[dict[str, Any]] = []
    review: list[dict[str, Any]] = []
    ref_order_map = {node.node_id: index for index, node in enumerate(ref_nodes, 1)}

    for node in nodes:
        if node.node_type != "SERIES":
            continue
        path = _path_for(node, by_id)
        ref = next((item for item in reversed(path) if item.node_type == "REF"), None)
        if ref is None:
            continue
        process = next((item for item in reversed(path) if item.node_type == "PROCESS"), None)
        code = _first_code(node.text)
        full_path = " > ".join(item.text for item in path)
        matches = ann_by_code.get(code, [])
        if not matches:
            matches = [None]
        for frame_index, record in enumerate(matches, 1):
            if record is None:
                symbol = symbol_raw = tolerance_display = tolerance_value = tolerance_unit = ""
                quantity_display = quantity_count = datum_references = raw_callout = ""
                source_method = "COM_DIRECT_NOT_FOUND"
                status = "NEEDS_REVIEW"
                review_needed = "Oui"
                review_reason = "annotation_code_not_found_in_AnnotationSets"
                record_id = ""
                review.append({
                    "record_id": "",
                    "reference_tree": ref.text,
                    "process": process.text if process else "",
                    "series_code": code,
                    "field_to_review": "annotation",
                    "value_found": "",
                    "raw_evidence": node.text,
                    "source_method": source_method,
                    "review_reason": review_reason,
                    "suggested_action": "Vérifier l'objet FTA correspondant dans CATIA.",
                    "review_status": "À traiter",
                    "validated_value": "",
                    "reviewer_note": "",
                    "capture_path": "",
                })
            else:
                symbol, symbol_raw = record.symbol, record.symbol_raw
                tolerance_display, tolerance_value, tolerance_unit = record.tolerance_display, record.tolerance_value, record.tolerance_unit
                quantity_display, quantity_count = record.quantity_display, record.quantity_count
                datum_references, raw_callout = record.datum_references, record.raw_callout
                source_method, status, review_needed, review_reason, record_id = (
                    record.source_method, record.status, record.review_needed, record.review_reason, record.record_id
                )
                if record.review_needed == "Oui":
                    review.append({
                        "record_id": record.record_id,
                        "reference_tree": ref.text,
                        "process": process.text if process else "",
                        "series_code": code,
                        "field_to_review": record.review_reason,
                        "value_found": raw_callout,
                        "raw_evidence": json.dumps(record.raw_fields, ensure_ascii=False, sort_keys=True),
                        "source_method": record.source_method,
                        "review_reason": record.review_reason,
                        "suggested_action": "Compléter depuis CATIA ou valider visuellement cette annotation.",
                        "review_status": "À traiter",
                        "validated_value": "",
                        "reviewer_note": "",
                        "capture_path": "",
                    })

            if not process:
                review.append({
                    "record_id": record_id,
                    "reference_tree": ref.text,
                    "process": "",
                    "series_code": code,
                    "field_to_review": "hierarchy_missing",
                    "value_found": node.text,
                    "raw_evidence": full_path,
                    "source_method": "COM_DIRECT",
                    "review_reason": "La série n'a pas de processus parent accessible.",
                    "suggested_action": "Vérifier la structure de l'arbre CATIA.",
                    "review_status": "À traiter",
                    "validated_value": "",
                    "reviewer_note": "",
                    "capture_path": "",
                })
                status = "NEEDS_REVIEW"
                review_needed = "Oui"
                review_reason = "; ".join(filter(None, (review_reason, "hierarchy_missing")))

            rows.append({
                "line": len(rows) + 1,
                "record_id": record_id,
                "ref_order": ref_order_map.get(ref.node_id, 0),
                "reference_tree": ref.text,
                "process_order": _node_order(process, siblings) if process else 0,
                "process_code": (re.match(r"\s*(\d{2})", process.text).group(1) if process and re.match(r"\s*(\d{2})", process.text) else ""),
                "process": process.text if process else "",
                "series_order": _node_order(node, siblings),
                "series_code": code,
                "series_label": node.text,
                "frame_index": frame_index,
                "symbol": symbol,
                "symbol_raw": symbol_raw,
                "tolerance_display": tolerance_display,
                "tolerance_value": tolerance_value,
                "tolerance_unit": tolerance_unit,
                "quantity_display": quantity_display,
                "quantity_count": quantity_count,
                "datum_references_frame": datum_references,
                "reference_context": ref.text,
                "raw_callout": raw_callout,
                "source_method": source_method,
                "status": status,
                "review_needed": review_needed,
                "review_reason": review_reason,
                "full_path": full_path,
            })
    return rows, review


def _tree_rows(nodes: Sequence[TreeNode]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    by_id = _ancestor_map(nodes)
    for line, node in enumerate(nodes, 1):
        path = " > ".join(item.text for item in _path_for(node, by_id))
        result.append({
            "line": line,
            "node_id": node.node_id,
            "parent_id": node.parent_id,
            "level": node.level,
            "node_type": node.node_type,
            "text": node.text,
            "full_path": path,
            "collection_source": node.collection_source,
            "collection_index": node.collection_index,
            "object_name": node.raw_properties.get("Name", ""),
            "object_type": node.object_type,
            "source_com_path": node.source_com_path,
            "children_count": node.children_count,
            "read_status": "OK" if not node.errors else "PARTIAL",
        })
    return result


def _serialisable(value: Any) -> Any:
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    if isinstance(value, (list, tuple)):
        return [_serialisable(item) for item in value]
    return _text(value)


def _write_csv(path: Path, rows: Sequence[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8-sig")
        return
    fieldnames: list[str] = []
    for row in rows:
        for key in row:
            if key not in fieldnames:
                fieldnames.append(key)
    with path.open("w", newline="", encoding="utf-8-sig") as stream:
        writer = csv.DictWriter(stream, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _format_worksheet(worksheet: Any) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    worksheet.freeze_panes = "A2"
    if worksheet.max_column and worksheet.max_row:
        worksheet.auto_filter.ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
    worksheet.row_dimensions[1].height = 30
    for column in range(1, worksheet.max_column + 1):
        letter = get_column_letter(column)
        header = worksheet.cell(1, column)
        header.fill = fill
        header.font = font
        header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        values = [header.value]
        for row in range(2, min(worksheet.max_row, 151) + 1):
            values.append(worksheet.cell(row, column).value)
        longest = max((len(str(value)) for value in values if value is not None), default=8)
        name = str(header.value or "").casefold()
        if name in {"full_path", "raw_callout", "raw_evidence", "properties", "source_com_path"}:
            width = min(72, max(28, longest + 2))
        elif name in {"text", "process", "series_label", "reference_tree", "review_reason", "suggested_action"}:
            width = min(48, max(16, longest + 2))
        else:
            width = min(25, max(10, longest + 2))
        worksheet.column_dimensions[letter].width = width


def _write_workbook(path: Path, sheets: dict[str, list[dict[str, Any]]]) -> Path:
    from openpyxl import Workbook

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    first = True
    for sheet_name, rows in sheets.items():
        worksheet = workbook.active if first else workbook.create_sheet()
        first = False
        worksheet.title = sheet_name[:31]
        fields: list[str] = []
        for row in rows:
            for key in row:
                if key not in fields:
                    fields.append(key)
        if not fields:
            fields = ["status"]
            rows = [{"status": "Aucune donnée"}]
        worksheet.append(fields)
        for row in rows:
            worksheet.append([_serialisable(row.get(field_name, "")) for field_name in fields])
        _format_worksheet(worksheet)
    try:
        workbook.save(path)
    except PermissionError:
        fallback = path.with_name(f"{path.stem}_{datetime.now().strftime('%H%M%S')}{path.suffix}")
        workbook.save(fallback)
        return fallback
    return path


def _new_run_directory(output_root: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    destination = output_root / timestamp
    suffix = 2
    while destination.exists():
        destination = output_root / f"{timestamp}_{suffix}"
        suffix += 1
    destination.mkdir(parents=True, exist_ok=False)
    return destination


def _connect_catia(document_path: Path | None) -> tuple[Any, Any]:
    try:
        import win32com.client  # type: ignore[import-not-found]
    except ImportError as error:
        raise FunctionalToleranceError(
            "pywin32 est nécessaire. Installez les dépendances du projet puis relancez le programme."
        ) from error
    # ``Dispatch`` can start a *new empty* CATIA process.  That is useful only
    # when the caller supplied a document path that we can open immediately;
    # otherwise it hides the real problem (a closed CATIA instance, or a
    # Windows privilege mismatch) behind the misleading "no document" error.
    attach_error: Exception | None = None
    try:
        catia = win32com.client.GetActiveObject("CATIA.Application")
    except Exception as error:
        attach_error = error
        # A CATIA instance may be registered under a ROT entry that pywin32's
        # GetActiveObject helper does not resolve.  GetObject(Class=...) is a
        # second, non-creating lookup; unlike Dispatch it cannot hide the
        # user's real CATIA session behind a new empty process.
        try:
            catia = win32com.client.GetObject(Class="CATIA.Application")
        except Exception as alternate_error:
            attach_error = alternate_error
            catia = None
        if document_path is None and catia is None:
            raise FunctionalToleranceError(
                "Impossible de se connecter à CATIA déjà ouvert. "
                "Ouvrez CATIA avec une pièce CATPart/CATProduct, puis relancez. "
                "Si CATIA est lancé en administrateur, lancez aussi PowerShell "
                "au même niveau de privilèges. Vous pouvez aussi utiliser "
                "--document \"C:\\chemin\\piece.CATPart\"."
            ) from error
        if catia is None:
            try:
                catia = win32com.client.Dispatch("CATIA.Application")
            except Exception as dispatch_error:
                raise FunctionalToleranceError(
                    "CATIA n'est pas accessible. Ouvrez CATIA ou fournissez un fichier "
                    "avec --document."
                ) from dispatch_error
    try:
        catia.Visible = True
    except Exception:
        pass

    # CATIA/pywin32 can expose ``Documents`` (and sometimes ``ActiveDocument``
    # or ``Count``) as either a property or a zero-argument COM method.  Read
    # only these known members through the bounded helper instead of rejecting
    # a callable COM wrapper as an error.
    documents, _, documents_error = _read_scalar_or_object(catia, "Documents")

    if document_path is not None:
        path = document_path.expanduser().resolve()
        if not path.is_file():
            raise FunctionalToleranceError(f"Document CATIA introuvable : {path}")
        current = ""
        active_document, _, _ = _read_scalar_or_object(catia, "ActiveDocument")
        if active_document is not None:
            current_value, _, _ = _read_scalar_or_object(active_document, "FullName")
            current = _text(current_value)
        if not current or Path(current).resolve() != path:
            if documents is None:
                raise FunctionalToleranceError(
                    "La collection Documents de CATIA est inaccessible : "
                    f"{documents_error or 'erreur COM inconnue'}"
                )
            opened, error = _safe_call(documents, "Open", str(path))
            if opened is None and error:
                raise FunctionalToleranceError(f"CATIA ne peut pas ouvrir {path} : {error}")

    document, _, active_error = _read_scalar_or_object(catia, "ActiveDocument")
    if document is None:
        # CATIA can return a fresh collection proxy after Documents.Open.
        documents, _, documents_error = _read_scalar_or_object(catia, "Documents")
        count, _, count_error = _read_scalar_or_object(documents, "Count")
        try:
            count_int = int(count or 0)
        except (TypeError, ValueError):
            count_int = 0
        if count_int < 1:
            detail = documents_error or count_error or active_error or attach_error
            raise FunctionalToleranceError(
                "CATIA est accessible, mais aucun document CATPart/CATProduct n'est ouvert. "
                "Ouvrez la pièce dans CATIA, attendez la fin du chargement, puis relancez. "
                "Pour ouvrir automatiquement une pièce, utilisez --document \"C:\\chemin\\piece.CATPart\"."
                + (f" Détail COM : {detail}" if detail else "")
            )
        document, item_error = _safe_call(documents, "Item", 1)
        if document is None and item_error:
            raise FunctionalToleranceError(
                f"CATIA contient {count_int} document(s), mais le premier est inaccessible : {item_error}"
            ) from item_error
    if document is None:
        raise FunctionalToleranceError("Le document CATIA actif est inaccessible.")
    return catia, document


def _load_tree_excel(path: Path) -> list[TreeNode]:
    """Optional hierarchy fallback for CATIA releases hiding UI groups."""
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise FunctionalToleranceError("openpyxl est nécessaire pour --tree-excel.") from error
    if not path.is_file():
        raise FunctionalToleranceError(f"Fichier d'arbre introuvable : {path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["CATIA_TREE"] if "CATIA_TREE" in workbook.sheetnames else workbook.active
    headers = [str(cell.value or "") for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    positions = {header: index for index, header in enumerate(headers)}
    required = {"text", "level"}
    if not required.issubset(positions):
        raise FunctionalToleranceError("La feuille d'arbre doit contenir au minimum les colonnes text et level.")
    nodes: list[TreeNode] = []
    stack: list[TreeNode] = []
    for line, cells in enumerate(sheet.iter_rows(min_row=2), 1):
        values = [cell.value for cell in cells]
        text_value = _clean(_text(values[positions["text"]]))
        try:
            level = int(values[positions["level"]] or 0)
        except (TypeError, ValueError):
            level = 0
        while stack and stack[-1].level >= level:
            stack.pop()
        parent = stack[-1] if stack else None
        node_id = _text(values[positions.get("node_id", -1)]) if positions.get("node_id", -1) >= 0 else f"X{line:05d}"
        node = TreeNode(
            node_id=node_id or f"X{line:05d}",
            parent_id=parent.node_id if parent else "",
            level=level,
            text=text_value,
            node_type=classify_node(text_value),
            collection_source="TREE_EXCEL",
            collection_index=line,
            source_com_path="TREE_EXCEL",
            object_type="",
            object_ref=None,
        )
        nodes.append(node)
        stack.append(node)
    return nodes


def _frame_change_ratio(previous: Any, current: Any) -> float:
    """Measure change in the left tree crop for the opt-in visual fallback."""
    if previous is None or current is None or getattr(previous, "shape", None) != getattr(current, "shape", None):
        return 1.0
    try:
        import cv2
        import numpy as np
    except ImportError:
        return 1.0
    height, width = previous.shape[:2]
    right = min(width, 520)
    left = min(30, max(0, right - 1))
    before = previous[:, left:right]
    after = current[:, left:right]
    difference = cv2.absdiff(before, after)
    gray = cv2.cvtColor(difference, cv2.COLOR_BGR2GRAY)
    changed = np.count_nonzero(gray > 18)
    return float(changed) / float(gray.size or 1)


def _visual_fallback_nodes(run_dir: Path, *, maximum_captures: int) -> tuple[list[TreeNode], Path]:
    """Capture only a manually positioned REF branch when COM lacks it.

    This fallback is deliberately opt-in.  It reuses the already validated
    project OCR pipeline, but it never touches ``main.py`` and it never OCRs
    tolerance values: FTA values still come from the direct COM reader.
    """
    try:
        import cv2  # noqa: F401
        import numpy as np  # noqa: F401
        from catia_tree_pipeline import extract_all_captures
        from annotation_text_recovery import recover_annotation_text
        from screen_capture import grab_catia_tree, save_tree_capture
        from tree_builder import TreeBuilder
        from tree_scroller import SCROLL_CLICKS, focus_tree, scroll_down
    except ImportError as error:
        raise FunctionalToleranceError(
            "Le secours visuel demande les modules OCR déjà présents dans le projet."
        ) from error

    capture_dir = run_dir / "visual_tree_captures"
    capture_dir.mkdir(parents=True, exist_ok=True)
    print("\n[SECours visuel] La hiérarchie REF n'est pas exposée par CATIA COM.")
    print("Placez manuellement le haut de la branche REF dans l'arbre CATIA.")
    input("Appuyez sur Entrée pour commencer les captures ciblées... ")
    first = grab_catia_tree(activate=False)
    if first is None or not save_tree_capture(first, capture_dir / "tree_000.png"):
        raise FunctionalToleranceError("Impossible de capturer la branche REF visible.")
    if not focus_tree():
        raise FunctionalToleranceError("CATIA n'a pas pu reprendre le focus de l'arbre.")

    previous = first
    stable_frames = 0
    scroll_clicks = SCROLL_CLICKS
    for index in range(1, max(2, int(maximum_captures))):
        if not scroll_down(focus=False, clicks=scroll_clicks):
            raise FunctionalToleranceError("Le défilement de l'arbre CATIA a échoué.")
        frame = grab_catia_tree(activate=False)
        if frame is None:
            raise FunctionalToleranceError("Une capture de la branche REF a échoué.")
        ratio = _frame_change_ratio(previous, frame)
        if ratio < 0.0008:
            stable_frames += 1
            if stable_frames >= 2:
                break
            scroll_clicks = min(180, max(45, int(round(scroll_clicks * 1.20))))
            continue
        stable_frames = 0
        if ratio < 0.018:
            scroll_clicks = min(180, max(45, int(round(scroll_clicks * 1.25))))
        elif ratio > 0.20:
            scroll_clicks = max(45, int(round(scroll_clicks * 0.75)))
        path = capture_dir / f"tree_{index:03d}.png"
        if not save_tree_capture(frame, path):
            raise FunctionalToleranceError(f"Impossible d'enregistrer la capture {index:03d}.")
        previous = frame
    else:
        raise FunctionalToleranceError(
            f"Le secours visuel a atteint la limite de {maximum_captures} captures."
        )

    raw = extract_all_captures(capture_dir)
    if raw is None or raw.empty:
        raise FunctionalToleranceError("Aucun libellé lisible dans les captures ciblées de la branche REF.")
    builder = TreeBuilder()
    builder.load_dataframe(raw)
    dataframe = recover_annotation_text(builder.build())
    nodes: list[TreeNode] = []
    for line, row in enumerate(dataframe.to_dict(orient="records"), 1):
        def optional(value: Any) -> str:
            rendered = _clean(_text(value))
            return "" if rendered.casefold() in {"nan", "none"} else rendered

        text_value = optional(row.get("corrected_text")) or optional(row.get("text"))
        node_id = optional(row.get("node_id")) or f"V{line:05d}"
        parent_id = optional(row.get("parent_id"))
        try:
            level = int(row.get("level", 0) or 0)
        except (TypeError, ValueError):
            level = 0
        node = TreeNode(
            node_id=node_id,
            parent_id=parent_id,
            level=level,
            text=text_value,
            node_type=classify_node(text_value),
            collection_source="VISUAL_FALLBACK",
            collection_index=line,
            source_com_path=optional(row.get("image")) or str(capture_dir),
            object_type="OCR",
            raw_properties={"ocr_conf": optional(row.get("conf"))},
            errors=["hierarchy_source_visual_fallback"],
        )
        nodes.append(node)
    return nodes, capture_dir


def _manifest(catia: Any, document: Any, *, nodes: Sequence[TreeNode], records: Sequence[AnnotationRecord], review: Sequence[dict[str, Any]], run_dir: Path, tree_source: str, truncated: bool) -> list[dict[str, Any]]:
    def scalar(owner: Any, prop: str) -> str:
        value, _, _ = _read_scalar_or_object(owner, prop)
        return _text(value)

    refs = sum(node.node_type == "REF" for node in nodes)
    processes = sum(node.node_type == "PROCESS" for node in nodes)
    series = sum(node.node_type == "SERIES" for node in nodes)
    direct = sum(record.source_method == "COM_DIRECT" for record in records)
    reviewed = sum(record.review_needed == "Oui" for record in records)
    return [
        {"field": "script_version", "value": SCRIPT_VERSION},
        {"field": "run_timestamp", "value": datetime.now().isoformat(timespec="seconds")},
        {"field": "document_name", "value": scalar(document, "Name")},
        {"field": "document_full_name", "value": scalar(document, "FullName")},
        {"field": "catia_version", "value": scalar(catia, "Version")},
        {"field": "tree_source", "value": tree_source},
        {"field": "output_run", "value": str(run_dir)},
        {"field": "nodes_read", "value": len(nodes)},
        {"field": "ref_count", "value": refs},
        {"field": "process_count", "value": processes},
        {"field": "series_count", "value": series},
        {"field": "fta_annotations_read", "value": len(records)},
        {"field": "com_direct_records", "value": direct},
        {"field": "records_needing_review", "value": reviewed},
        {"field": "review_rows", "value": len(review)},
        {"field": "traversal_truncated", "value": "Oui" if truncated else "Non"},
    ]


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Lecture directe CATIA des branches REF et des tolérances fonctionnelles.")
    parser.add_argument("--document", type=Path, help="CATPart/CATProduct à ouvrir ; sinon le document CATIA actif est utilisé.")
    parser.add_argument("--output-dir", type=Path, help="Dossier racine des résultats (par défaut : results/functional_tolerances/runs).")
    parser.add_argument("--tree-excel", type=Path, help="Excel d'arbre optionnel si la hiérarchie REF n'est pas exposée par COM.")
    parser.add_argument(
        "--visual-fallback",
        action="store_true",
        help="Si COM ne trouve aucune branche REF, demander une capture manuelle ciblée de cette branche.",
    )
    parser.add_argument("--max-captures", type=int, default=250, help="Limite du secours visuel ciblé.")
    parser.add_argument("--diagnose", action="store_true", help="Afficher et conserver davantage d'informations COM.")
    parser.add_argument("--max-depth", type=int, default=DEFAULT_MAX_DEPTH)
    parser.add_argument("--max-nodes", type=int, default=DEFAULT_MAX_NODES)
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    if args.max_depth < 1 or args.max_nodes < 1:
        print("[ERROR] --max-depth et --max-nodes doivent être positifs.")
        return 2
    if args.max_captures < 2:
        print("[ERROR] --max-captures doit être au moins égal à 2.")
        return 2
    project_root = Path(__file__).resolve().parent
    output_root = (args.output_dir or project_root / "results" / "functional_tolerances" / "runs").expanduser().resolve()
    run_dir = _new_run_directory(output_root)
    try:
        print("=" * 72)
        print("CATIA FUNCTIONAL TOLERANCE - DIRECT COM EXPORT")
        print("=" * 72)
        print("Lecture directe du modèle CATIA ; le programme de captures main.py n'est pas utilisé.")
        catia, document = _connect_catia(args.document)

        model_reader = CatiaModelReader(max_depth=args.max_depth, max_nodes=args.max_nodes)
        com_nodes = model_reader.read(document)
        tree_source = "CATIA_COM"
        nodes = com_nodes
        if args.tree_excel is not None:
            fallback_nodes = _load_tree_excel(args.tree_excel.expanduser().resolve())
            if sum(node.node_type == "REF" for node in fallback_nodes) > sum(node.node_type == "REF" for node in nodes):
                nodes = fallback_nodes
                tree_source = "CATIA_COM_PLUS_TREE_EXCEL"

        if args.visual_fallback and not any(node.node_type == "REF" for node in nodes):
            visual_nodes, capture_dir = _visual_fallback_nodes(run_dir, maximum_captures=args.max_captures)
            if any(node.node_type == "REF" for node in visual_nodes):
                nodes = visual_nodes
                tree_source = f"CATIA_COM_PLUS_VISUAL_FALLBACK:{capture_dir}"

        fta_reader = FtaReader()
        records, audit_rows = fta_reader.read(document)
        tolerance_rows, review_rows = build_tolerance_rows(nodes, records)
        if not any(node.node_type == "REF" for node in nodes):
            review_rows.append({
                "record_id": "",
                "reference_tree": "",
                "process": "",
                "series_code": "",
                "field_to_review": "hierarchy",
                "value_found": "",
                "raw_evidence": " | ".join(node.text for node in nodes[:50]),
                "source_method": tree_source,
                "review_reason": "Aucune branche REF n'a été exposée par le modèle COM.",
                "suggested_action": "Relancer avec --tree-excel ou fournir une capture ciblée de la branche REF.",
                "review_status": "À traiter",
                "validated_value": "",
                "reviewer_note": "",
                "capture_path": "",
            })

        sheets = {
            "TOLERANCES": tolerance_rows,
            "ARBRE_FONCTIONNEL": _tree_rows(nodes),
            "LECTURE_BRUTE": audit_rows,
            "REVIEW": review_rows,
            "MANIFESTE": _manifest(
                catia,
                document,
                nodes=nodes,
                records=records,
                review=review_rows,
                run_dir=run_dir,
                tree_source=tree_source,
                truncated=model_reader._truncated,
            ),
        }
        csv_dir = run_dir / "csv"
        json_dir = run_dir / "json"
        excel_dir = run_dir / "excel"
        for sheet_name, rows in sheets.items():
            _write_csv(csv_dir / f"{sheet_name.casefold()}.csv", rows)
            json_dir.mkdir(parents=True, exist_ok=True)
            (json_dir / f"{sheet_name.casefold()}.json").write_text(
                json.dumps([_serialisable(row) for row in rows], ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
        workbook_path = _write_workbook(excel_dir / "functional_tolerances.xlsx", sheets)
        (run_dir / "manifest.json").write_text(
            json.dumps(_manifest(catia, document, nodes=nodes, records=records, review=review_rows, run_dir=run_dir, tree_source=tree_source, truncated=model_reader._truncated), ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

        print(f"[OK] Branches REF : {sum(node.node_type == 'REF' for node in nodes)}")
        print(f"[OK] Processus    : {sum(node.node_type == 'PROCESS' for node in nodes)}")
        print(f"[OK] Séries       : {sum(node.node_type == 'SERIES' for node in nodes)}")
        print(f"[OK] Annotations FTA lues : {len(records)}")
        print(f"[OK] À vérifier   : {len(review_rows)}")
        print(f"Excel : {workbook_path}")
        if args.diagnose:
            print(f"Audit COM : {run_dir / 'json' / 'lecture_brute.json'}")
        print(f"Résultats : {run_dir}")
        return 0
    except FunctionalToleranceError as error:
        print(f"[ERROR] {error}")
        if args.diagnose:
            (run_dir / "error.txt").write_text(str(error), encoding="utf-8")
        print(f"Le dossier de diagnostic reste disponible : {run_dir}")
        return 2
    except Exception as error:
        print(f"[ERROR] Erreur inattendue : {_exception_text(error)}")
        if args.diagnose:
            (run_dir / "error.txt").write_text(_exception_text(error), encoding="utf-8")
        return 3


if __name__ == "__main__":
    raise SystemExit(main())
